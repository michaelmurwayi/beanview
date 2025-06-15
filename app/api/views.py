from django.shortcuts import render
from .models import *
from .serializers import *
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db.models import Sum
from datetime import datetime, timedelta
import json
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from rest_framework import status
from rest_framework.response import Response
from .process_mill_statements import DataCleaner
from .coffee.read_file import read_xls_file
from .coffee.clean_masterlog_df import clean_outturns
from .coffee.check_pockets import check_for_pockets
import os
import json 
from .process_records.record_processing import process_uploaded_files, process_single_record
import csv
from io import StringIO
from rest_framework.exceptions import ValidationError
from django.conf import settings
from openpyxl import load_workbook
import traceback
from openpyxl.cell.cell import MergedCell
import pandas as pd
from copy import copy





class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer

@method_decorator(csrf_exempt, name='dispatch')
class FarmersViewSet(viewsets.ModelViewSet):
    queryset = Farmer.objects.all()
    serializer_class = FarmerSerializer

    def create(self, request):
        data = request.data.dict()
        # Serialize the combined data
        
        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)

        # Save the instance if valid
        self.perform_create(serializer)
        # Use the saved instance for headers
        headers = self.get_success_headers(serializer.instance)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)
        
@method_decorator(csrf_exempt, name='dispatch')
class CoffeeViewSet(viewsets.ModelViewSet):
    queryset = Coffee.objects.all()
    serializer_class = CoffeeSerializer

    def create(self, request, *args, **kwargs):
        data = request.data.dict() if hasattr(request.data, 'dict') else request.data
        files = request.FILES
        sheets = data.get("sheetnames", "").split(",") if data.get("sheetnames") else []

        if files and sheets:
            return process_uploaded_files(self, data, sheets)
        return process_single_record(self, data)

    def update(self, request, *args, **kwargs):
        """Handle the PUT method for updating a Coffee record."""
        
        instance = self.get_object()
        # PUT should usually update entire resource, so partial=False
        serializer = self.get_serializer(instance, data=request.data, partial=True)
        
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def perform_update(self, serializer):
        serializer.save()

    @action(detail=False, methods=['GET'], url_path='total_net_weight')
    def total_net_weight(self, request):
        try:
            # Exclude coffee with status 'SOLD' (compare via related CoffeeStatus.name)
            records = self.queryset.exclude(status_id=1).values_list('weight', flat=True)
            total_net_weight = sum(records)
            return Response({"total_net_weight": total_net_weight})
        except Exception as e:
            return Response(
                {"error": f"Error calculating total net weight: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST
            )
    @action(detail=False, methods=['GET'], url_path='total_tare_weight')
    def total_tare_weight(self, request):
        try:
            records = self.queryset.exclude(status="SOLD").values_list('tare_weight', flat=True)
            total_tare_weight = sum(records)
            return Response({"total_tare_weight": total_tare_weight})
        except Exception as e:
            return Response({"error": f"Error calculating total tare weight: {str(e)}"}, status=400)

    @action(detail=False, methods=['GET'], url_path='total_number_bags')
    def total_number_bags(self, request):
        try:
            records = self.queryset.exclude(status="SOLD").values_list('bags', flat=True)
            total_number_bags = sum(records)
            return Response({"total_number_bags": total_number_bags})
        except Exception as e:
            return Response({"error": f"Error calculating total number of bags: {str(e)}"}, status=400)

    @action(detail=False, methods=['GET'], url_path='total_number_farmers')
    def total_number_farmers(self, request):
        try:
            total_number_farmers = self.queryset.values("estate").distinct().count()
            return Response({"total_number_farmers": total_number_farmers})
        except Exception as e:
            return Response({"error": f"Error calculating total number of farmers: {str(e)}"}, status=400)

    @action(detail=False, methods=['GET'], url_path='performance_per_grade')
    def performance_per_grade(self, request):
        try:
            performance_per_grade = []
            distinct_grades = self.queryset.values("grade").distinct()
            for record in distinct_grades:
                grade = record["grade"]
                grade_records = self.queryset.filter(grade=grade)
                total_weight = grade_records.aggregate(total_weight=Sum('net_weight'))['total_weight'] or 0
                performance_per_grade.append({"grade": grade, "net_weight": total_weight})
            return Response({"Performance": performance_per_grade})
        except Exception as e:
            return Response({"error": f"Error calculating performance per grade: {str(e)}"}, status=400)

    @action(detail=False, methods=['GET'], url_path='daily_delivery')
    def daily_delivery(self, request):
        try:
            deliveries = []
            records = self.queryset.values()
            for record in records:
                target_date = record.get('created_at')
                if target_date and is_less_than_24_hours_ago(target_date):
                    deliveries.append(record)
            return Response({"deliveries": deliveries})
        except Exception as e:
            return Response({"error": f"Error fetching daily deliveries: {str(e)}"}, status=400)

    @action(detail=False, methods=['POST'])
    def generate_summary_file(self, request, *args, **kwargs):
        TEMPLATE_PATH = os.path.join(settings.MEDIA_ROOT, 'templates', 'sale summary template.xlsx')
        START_ROW = 27

        try:
            summaries = request.data.get('summaries', [])

            if not summaries:
                raise ValidationError("'summaries' is required and must not be empty.")

            base_dir = os.path.join(settings.MEDIA_ROOT, 'summaries')
            os.makedirs(base_dir, exist_ok=True)

            # Collect all status IDs from all records
            all_status_ids = {
                record.get('status_id')
                for summary in summaries
                for record in summary.get('records', [])
                if record.get('status_id') is not None
            }

            status_map = {
                status.id: status.name
                for status in CoffeeStatus.objects.filter(id__in=all_status_ids)
            }

            generated_files = []

            for summary in summaries:
                mark = summary.get('mark')
                records = summary.get('records', [])

                if not mark or not records:
                    continue  # skip invalid

                total_weight = sum(float(r.get('weight') or 0) for r in records)
                count = len(records)

                mark_dir = os.path.join(base_dir, mark)
                os.makedirs(mark_dir, exist_ok=True)

                filename = f"{mark}_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                file_path = os.path.join(mark_dir, filename)

                wb = load_workbook(TEMPLATE_PATH)
                ws = wb.active

                # Write summary data
               
                ws['B6'] = mark
               
                for row_offset, record in enumerate(records, start=1):
                    row = START_ROW + row_offset
                    status_id = record.get('status')
                    status_name = CoffeeStatus.objects.get(id=status_id).name

                    values = [
                        record.get('outturn'),
                        record.get('bulkoutturn'),
                        record.get('mark'),
                        record.get('type'),
                        record.get('grade'),
                        record.get('bags'),
                        record.get('pockets'),
                        record.get('weight'),
                        record.get('sale_number'),
                        record.get('season'),
                        record.get('certificate'),
                        record.get('mill'),
                        record.get('warehouse'),
                        record.get('price'),
                        record.get('buyer'),
                        status_name,
                    ]

                    for col_index, value in enumerate(values, start=1):
                        cell = ws.cell(row=row, column=col_index)
                        if isinstance(cell, MergedCell):
                            continue
                        cell.value = value

                wb.save(file_path)
                generated_files.append({"mark": mark, "path": file_path})

            return Response({"message": "Files generated", "files": generated_files}, status=200)

        except ValidationError as e:
            traceback.print_exc()
            return Response({"error": str(e)}, status=400)

        except Exception as e:
            traceback.print_exc()
            return Response({"error": f"Internal server error: {str(e)}"}, status=500)


def assign_lots(df, start_lot=7301):
    df = df.copy()
    df["LOT"] = list(range(start_lot, start_lot + len(df)))
    return df, len(df)


def summarize_grades(df):
    summary = df.groupby("grade")["bags"].sum().to_dict()
    total_bags = df["bags"].sum()
    return summary, total_bags

def write_grade_summary(ws, summary, start_row=19, start_col=8):
    row = start_row
    for grade, bags in summary.items():
        # Write the grade in the specified column
        cell_grade = ws.cell(row=row, column=start_col)
        # Write the bags in the column to the right
        cell_bags = ws.cell(row=row, column=start_col + 1)
        cell_grade.value = grade
        cell_bags.value = bags
        row += 1  # Move to the next row for the next pair


def write_summary_to_excel(ws, num_bags, num_lots):
    summary_text = f"{num_bags} bags of Kenya Coffee In {num_lots} Lots"
    ws["I9"] = summary_text

def write_warehouse_location(ws, records):
    # Extract unique warehouse IDs from records
    warehouse_ids = {record.get("warehouse") for record in records if record.get("warehouse")}
    
    # Get warehouse names from DB
    warehouses = Warehouse.objects.filter(id__in=warehouse_ids).values_list("name", flat=True)
    warehouses = sorted({w.strip().upper() for w in warehouses if w})

    # Format location string
    if not warehouses:
        location_text = "Located at (No warehouse info)"
    elif len(warehouses) == 1:
        location_text = f"Located at ({warehouses[0]} warehouse)"
    else:
        location_text = f"Located at ({' & '.join(warehouses)} warehouse)"

    ws["I10"] = location_text

def write_milled_by(ws, records, start_row=12, column_letter="A"):
    # Get unique mill IDs from the records
    mill_ids = {record.get("mill") for record in records if record.get("mill")}
    
    # Query the Mill model for names and codes
    mills = Mill.objects.filter(id__in=mill_ids).values_list("name", "location")

    # Clean, deduplicate, and sort
    unique_mills = sorted({(name.strip(), code.strip()) for name, code in mills if name and code})

    # Write each mill on its own line starting from `start_row`
    for i, (name, code) in enumerate(unique_mills):
        text = f"Milled BY({name} Coffee Mill Denoted as {code})"
        cell = f"{column_letter}{start_row + i}"
        ws[cell] = text

def replace_mill_ids_with_names(df):
    if "mill" not in df.columns:
        return df  # nothing to do

    # Get unique mill IDs from the DataFrame
    mill_ids = df["mill"].dropna().unique().tolist()

    # Fetch mill names from the database
    mill_map = dict(Mill.objects.filter(id__in=mill_ids).values_list("id", "name"))

    # Replace the IDs with mill names in the DataFrame
    df["mill"] = df["mill"].map(mill_map).fillna("")

    return df
def replace_warehouse_ids_with_names(df):
    if "warehouse" not in df.columns:
        return df
    # Get unique warehouse IDs from the DataFrame
    warehouse_ids = df["warehouse"].dropna().unique().tolist()
    # Fetch warehouse names from the database
    warehouse_map = dict(Warehouse.objects.filter(id__in=warehouse_ids).values_list("id", "name"))
    # Replace the IDs with warehouse names in the DataFrame
    df["warehouse"] = df["warehouse"].map(warehouse_map).fillna("")

    return df

class CatalogueViewSet(viewsets.ModelViewSet):
    queryset = Catalogue.objects.all()
    serializer_class = CatalogueSerializer

    @action(detail=False, methods=['POST'])
    def generate_catalogue_file(self, request):
        try:
            sale_number = request.data.get("sale")
            records = request.data.get("records")

            if not sale_number or not records:
                return Response(
                    {"error": "Sale number and records are required."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            if not isinstance(records, list):
                return Response({"error": "Records must be a list."}, status=400)

            df = pd.DataFrame(records)
            df, num_lots = assign_lots(df)
            df = replace_mill_ids_with_names(df)
            df = replace_warehouse_ids_with_names(df)
            grade_summary, num_bags = summarize_grades(df)

            column_map = {
                "LOT": "LOT",
                "C_OUTTURN": "outturn",
                "MARK": "mark",
                "GRADE": "grade",
                "BAGS": "bags",
                "POCKETS": "pockets",
                "Weight": "weight",
                "SALE NO": "sale",
                "SEASON": "season",
                "CERTIFICATE": "certificate",
                "MILL": "mill",
                "W/H": "warehouse",
                "AGENT CODE": "agentCode",
                "RESERVE PRICE": "reserve",
                "REMARKS": ""  # optional or blank
            }

            data_for_template = {}
            for template_col, data_col in column_map.items():
                if data_col and data_col in df.columns:
                    data_for_template[template_col] = df[data_col]
                else:
                    data_for_template[template_col] = [""] * len(df)

            df_template = pd.DataFrame(data_for_template)

            if df_template.empty:
                return Response(
                    {"error": "No valid data found in provided records after mapping."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            template_path = os.path.join(settings.BASE_DIR, "media/templates", "catalogue_template.xlsx")
            wb = load_workbook(template_path)
            ws = wb.active

            start_row = 49
            template_row = 49

            for i, row in df_template.iterrows():
                for j, col in enumerate(column_map.keys()):
                    target_cell = ws.cell(row=start_row + i, column=j + 1)
                    template_cell = ws.cell(row=template_row, column=j + 1)

                    target_cell.value = row.get(col, "")
                    if template_cell.has_style:
                        target_cell.font = copy(template_cell.font)
                        target_cell.border = copy(template_cell.border)
                        target_cell.fill = copy(template_cell.fill)
                        target_cell.number_format = copy(template_cell.number_format)
                        target_cell.protection = copy(template_cell.protection)
                        target_cell.alignment = copy(template_cell.alignment)

            # ✅ Write summaries
            write_summary_to_excel(ws, num_bags, num_lots)
            write_warehouse_location(ws, records)
            write_milled_by(ws, records, start_row=11, column_letter="H")
            write_grade_summary(ws, grade_summary, start_row=19, start_col=8)


            subdir = str(sale_number)
            dir_path = os.path.join(settings.MEDIA_ROOT, "catalogue", subdir)
            os.makedirs(dir_path, exist_ok=True)

            filename = "catalogue_file.xlsx"
            filepath = os.path.join(dir_path, filename)
            wb.save(filepath)

            file_url = os.path.join(settings.MEDIA_URL, "catalogue", subdir, filename)

            return Response({
                "message": f"Catalogue file created for sale {sale_number}.",
                "file_url": file_url
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            traceback_str = traceback.format_exc()
            print(traceback_str)
            return Response(
                {"error": "Internal server error.", "details": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

def is_less_than_24_hours_ago(target_date):
    # Get the current date and time
    current_date = datetime.now()

    # Calculate the difference between the current date and the target date
    time_difference = current_date - target_date

    # Check if the difference is less than 24 hours
    return time_difference < timedelta(hours=24)

def read_data_from_pdf_file(mill,file,requests):
    data = []
    import ipdb;ipdb.set_trace()

    return data

